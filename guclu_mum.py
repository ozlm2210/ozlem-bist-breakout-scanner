"""BIST Güçlü Mum - Sade Takip v2, Streamlit sürümü.

Kullanıcının Colab kodundaki güçlü mum koşulları korunur.
Çıktılar yalnız dört bölümden oluşur:
- Günlük Güçlü Mum
- Haftalık Güçlü Mum
- Aylık Güçlü Mum
- Takip Listesi
"""

from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
import math

import numpy as np
import pandas as pd
import streamlit as st
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Güçlü mum şartları — gönderilen Colab koduyla aynıdır.
MIN_BREAKOUT_PCT = 3.0
MIN_RANGE_EXPANSION = 1.20
MIN_BODY_RATIO = 0.50
MIN_CLOSE_POS = 0.60
MIN_RSI = 50.0
MIN_VOL_RATIO = 0.80
RSI_LEN = 14
LOOKBACK_DAILY = 10

# Bozulma şartları — takip sayfası için.
SERT_DUSUS_ESIK = -7.0
KOTU_KAPANIS_POZ = 0.25
HACIMLI_SATIS_ORAN = 1.50


@dataclass(frozen=True)
class GucluMumResult:
    gunluk: pd.DataFrame
    haftalik: pd.DataFrame
    aylik: pd.DataFrame
    takip: pd.DataFrame
    failed: pd.DataFrame


def _fix_columns(df: pd.DataFrame | None) -> pd.DataFrame | None:
    if df is None or df.empty:
        return None
    out = df.copy()
    if isinstance(out.columns, pd.MultiIndex):
        out.columns = out.columns.get_level_values(0)
    out.columns = [str(c).strip().capitalize() for c in out.columns]
    if not all(c in out.columns for c in ["Open", "High", "Low", "Close"]):
        return None
    if "Volume" not in out.columns:
        out["Volume"] = np.nan
    out = out[["Open", "High", "Low", "Close", "Volume"]]
    out = out.replace([np.inf, -np.inf], np.nan)
    out = out.dropna(subset=["Open", "High", "Low", "Close"]).sort_index()
    if isinstance(out.index, pd.DatetimeIndex) and out.index.tz is not None:
        out.index = out.index.tz_localize(None)
    return out if not out.empty else None


def _download_daily(symbol: str) -> pd.DataFrame | None:
    try:
        raw = yf.download(
            f"{symbol}.IS",
            period="15y",
            interval="1d",
            auto_adjust=False,
            progress=False,
            threads=False,
        )
        return _fix_columns(raw)
    except Exception:
        return None


def _resample(df: pd.DataFrame, rule: str) -> pd.DataFrame | None:
    out = df.resample(rule).agg(
        {"Open": "first", "High": "max", "Low": "min", "Close": "last", "Volume": "sum"}
    )
    out = out.dropna(subset=["Open", "High", "Low", "Close"])
    return out if not out.empty else None


def _rsi(close: pd.Series, length: int = 14) -> pd.Series:
    delta = close.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    avg_gain = gain.ewm(alpha=1 / length, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1 / length, adjust=False).mean()
    rs = avg_gain / avg_loss.replace(0, np.nan)
    return 100 - (100 / (1 + rs))


def _ema(series: pd.Series, length: int) -> pd.Series:
    return series.ewm(span=length, adjust=False).mean()


def _metrics(df: pd.DataFrame) -> pd.DataFrame:
    d = df.copy()
    d["Body"] = (d["Close"] - d["Open"]).abs()
    d["Range"] = d["High"] - d["Low"]
    d["RSI14"] = _rsi(d["Close"], RSI_LEN)
    d["EMA10"] = _ema(d["Close"], 10)
    d["EMA20"] = _ema(d["Close"], 20)
    d["Degisim_%"] = d["Close"].pct_change() * 100
    return d


def _num(value, digits=2):
    try:
        value = float(value)
        if np.isnan(value) or np.isinf(value):
            return None
        return round(value, digits)
    except Exception:
        return None


def _strong_candle(df: pd.DataFrame | None, symbol: str, period: str, lookback: int, last_trade: str):
    if df is None or len(df) < max(lookback + 1, 30):
        return None
    d = _metrics(df)
    last = d.iloc[-1]
    prev = d.iloc[-(lookback + 1):-1]

    candle_range = float(last["Range"])
    if not np.isfinite(candle_range) or candle_range <= 0:
        return None

    prev_high = float(prev["High"].max())
    avg_range = float(prev["Range"].mean())
    avg_volume = float(prev["Volume"].mean())
    high = float(last["High"])
    open_ = float(last["Open"])
    low = float(last["Low"])
    close = float(last["Close"])
    volume = float(last["Volume"])
    body = float(last["Body"])
    rsi14 = float(last["RSI14"])

    breakout = ((high / prev_high) - 1) * 100 if prev_high > 0 else np.nan
    range_exp = candle_range / avg_range if avg_range > 0 else np.nan
    body_ratio = body / candle_range
    close_pos = (close - low) / candle_range
    vol_ratio = volume / avg_volume if avg_volume > 0 else np.nan

    passed = (
        np.isfinite(breakout) and breakout >= MIN_BREAKOUT_PCT
        and np.isfinite(range_exp) and range_exp >= MIN_RANGE_EXPANSION
        and np.isfinite(body_ratio) and body_ratio >= MIN_BODY_RATIO
        and np.isfinite(rsi14) and rsi14 >= MIN_RSI
        and close > open_
        and np.isfinite(close_pos) and close_pos >= MIN_CLOSE_POS
        and np.isfinite(vol_ratio) and vol_ratio >= MIN_VOL_RATIO
    )
    if not passed:
        return None

    score = 0.0
    score += min(max(breakout, -20), 40) * 3.0
    score += min(max(range_exp, 0), 5) * 18
    score += min(max(body_ratio, 0), 1) * 20
    score += min(max(rsi14 - 45, 0), 35)
    score += min(max(vol_ratio, 0), 4) * 4
    score += 8
    score += close_pos * 10

    return {
        "Hisse": symbol,
        "Periyot": period,
        "Mum_Tarihi": pd.Timestamp(d.index[-1]).strftime("%Y-%m-%d"),
        "Son_Islem_Tarihi": last_trade,
        "Acilis": _num(open_, 4),
        "Yuksek": _num(high, 4),
        "Dusuk": _num(low, 4),
        "Kapanis": _num(close, 4),
        "Onceki_Tepe": _num(prev_high, 4),
        "Tepe_Ustu_%": _num(breakout),
        "Range_Genisleme": _num(range_exp),
        "Govde_Orani": _num(body_ratio, 4),
        "Kapanis_Pozisyonu": _num(close_pos, 4),
        "Hacim_Oran": _num(vol_ratio),
        "RSI14": _num(rsi14),
        "Skor": _num(score),
        "Yesil_Mum": "EVET",
        "Lookback": lookback,
    }


def _bozulma(df: pd.DataFrame) -> dict:
    d = _metrics(df)
    last = d.iloc[-1]
    prev = d.iloc[-11:-1]
    close, open_ = float(last["Close"]), float(last["Open"])
    high, low = float(last["High"]), float(last["Low"])
    rng = high - low
    close_pos = (close - low) / rng if rng > 0 else np.nan
    avg_vol = float(prev["Volume"].mean())
    vol_ratio = float(last["Volume"]) / avg_vol if avg_vol > 0 else np.nan
    degisim = float(last["Degisim_%"])
    ema10, ema20 = float(last["EMA10"]), float(last["EMA20"])
    red = close < open_
    reasons = []
    if np.isfinite(degisim) and degisim <= SERT_DUSUS_ESIK:
        reasons.append(f"Sert düşüş: {degisim:.2f}%")
    if np.isfinite(close_pos) and close_pos <= KOTU_KAPANIS_POZ:
        reasons.append(f"Kapanış dipte: {close_pos:.4f}")
    if np.isfinite(ema10) and close < ema10:
        reasons.append("EMA10 altı")
    if red and np.isfinite(vol_ratio) and vol_ratio >= HACIMLI_SATIS_ORAN:
        reasons.append(f"Hacimli kırmızı mum: {vol_ratio:.2f}x")
    return {
        "Son_Fiyat": _num(close, 4),
        "Son_Gun_Degisim_%": _num(degisim),
        "Gunluk_RSI14": _num(last["RSI14"]),
        "Gunluk_EMA10": _num(ema10, 4),
        "Gunluk_EMA20": _num(ema20, 4),
        "Bozuldu": "EVET" if reasons else "HAYIR",
        "Bozulma_Sebebi": " | ".join(reasons),
    }


def _scan_one(symbol: str, weekly_lookback: int, monthly_lookback: int):
    daily = _download_daily(symbol)
    if daily is None or len(daily) < 80:
        return symbol, None, "Veri alınamadı veya yetersiz"
    weekly = _resample(daily, "W-FRI")
    monthly = _resample(daily, "ME")
    last_trade = pd.Timestamp(daily.index[-1]).strftime("%Y-%m-%d")
    rows = [
        _strong_candle(daily, symbol, "Gunluk", LOOKBACK_DAILY, last_trade),
        _strong_candle(weekly, symbol, "Haftalik", weekly_lookback, last_trade),
        _strong_candle(monthly, symbol, "Aylik", monthly_lookback, last_trade),
    ]
    return symbol, (rows, _bozulma(daily)), ""


def run_guclu_mum(symbols: list[str], progress=None, workers: int = 6) -> GucluMumResult:
    today = datetime.now()
    year_start = datetime(today.year, 1, 1)
    weekly_lookback = max(1, math.ceil((today - year_start).days / 7))
    monthly_lookback = today.month

    all_rows, latest_map, failed = [], {}, []
    total = max(len(symbols), 1)
    with ThreadPoolExecutor(max_workers=max(1, workers)) as pool:
        futures = {pool.submit(_scan_one, s, weekly_lookback, monthly_lookback): s for s in symbols}
        for idx, future in enumerate(as_completed(futures), 1):
            symbol = futures[future]
            try:
                _, payload, error = future.result()
                if error or payload is None:
                    failed.append({"Hisse": symbol, "Sebep": error or "Bilinmeyen hata"})
                else:
                    rows, latest = payload
                    all_rows.extend([r for r in rows if r is not None])
                    latest_map[symbol] = latest
            except Exception as exc:
                failed.append({"Hisse": symbol, "Sebep": str(exc)})
            if progress:
                progress(idx / total, f"{idx}/{len(symbols)} tarandı")

    all_df = pd.DataFrame(all_rows)
    cols = ["Hisse", "Periyot", "Mum_Tarihi", "Son_Islem_Tarihi", "Acilis", "Yuksek", "Dusuk", "Kapanis", "Onceki_Tepe", "Tepe_Ustu_%", "Range_Genisleme", "Govde_Orani", "Kapanis_Pozisyonu", "Hacim_Oran", "RSI14", "Skor", "Yesil_Mum", "Lookback"]
    if all_df.empty:
        all_df = pd.DataFrame(columns=cols)

    def period_frame(name):
        out = all_df[all_df["Periyot"] == name].copy()
        return out.sort_values(["Skor", "Tepe_Ustu_%", "Hacim_Oran"], ascending=False).reset_index(drop=True)

    gunluk = period_frame("Gunluk")
    haftalik = period_frame("Haftalik")
    aylik = period_frame("Aylik")

    tracking = []
    if not all_df.empty:
        for symbol, group in all_df.groupby("Hisse"):
            periods = group["Periyot"].tolist()
            latest = latest_map.get(symbol, {})
            tracking.append({
                "Hisse": symbol,
                "Gunluk": "EVET" if "Gunluk" in periods else "HAYIR",
                "Haftalik": "EVET" if "Haftalik" in periods else "HAYIR",
                "Aylik": "EVET" if "Aylik" in periods else "HAYIR",
                "Periyot_Sayisi": len(set(periods)),
                "Periyotlar": ", ".join(sorted(set(periods))),
                "Ortalama_Skor": _num(group["Skor"].mean()),
                "En_Yuksek_Skor": _num(group["Skor"].max()),
                "En_Yuksek_Hacim_Oran": _num(group["Hacim_Oran"].max()),
                "Ortalama_RSI14": _num(group["RSI14"].mean()),
                "Son_Islem_Tarihi": group["Son_Islem_Tarihi"].iloc[0],
                **latest,
            })
    takip = pd.DataFrame(tracking)
    if not takip.empty:
        takip = takip.sort_values(["Bozuldu", "Periyot_Sayisi", "Ortalama_Skor"], ascending=[True, False, False]).reset_index(drop=True)
    return GucluMumResult(gunluk, haftalik, aylik, takip, pd.DataFrame(failed))


def _excel_bytes(result: GucluMumResult) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        result.gunluk.to_excel(writer, "GUNLUK_GUCLU_MUM", index=False)
        result.haftalik.to_excel(writer, "HAFTALIK_GUCLU_MUM", index=False)
        result.aylik.to_excel(writer, "AYLIK_GUCLU_MUM", index=False)
        result.takip.to_excel(writer, "TAKIP_LISTESI", index=False)
    output.seek(0)
    wb = load_workbook(output)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill, cell.font = header_fill, header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value == "EVET": cell.fill = green
                elif cell.value == "HAYIR": cell.fill = red
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        headers = [c.value for c in ws[1]]
        for name in ["Skor", "Ortalama_Skor", "En_Yuksek_Skor"]:
            if name in headers:
                col = headers.index(name) + 1
                for r in range(2, ws.max_row + 1):
                    try:
                        v = float(ws.cell(r, col).value)
                        if v >= 90: ws.cell(r, col).fill = green
                        elif v >= 70: ws.cell(r, col).fill = yellow
                    except Exception: pass
        for column in ws.columns:
            letter = get_column_letter(column[0].column)
            width = max((len(str(c.value)) if c.value is not None else 0) for c in column) + 3
            ws.column_dimensions[letter].width = min(width, 45)
    final = BytesIO()
    wb.save(final)
    return final.getvalue()


def render_guclu_mum(symbols: list[str]) -> None:
    st.title("BIST Güçlü Mum – Sade Takip v2")
    st.caption("Gönderilen Colab kodundaki güçlü mum koşulları değiştirilmeden uygulanır.")
    st.info("Çıktı: Günlük, Haftalık, Aylık Güçlü Mum ve birleşik Takip Listesi.")

    c1, c2 = st.columns(2)
    with c1:
        workers = st.slider("Paralel işlem", 2, 10, 6, 1, key="guclu_workers")
    with c2:
        st.metric("Taranacak hisse", len(symbols))

    if st.button("Güçlü Mum Taramasını Başlat", type="primary", use_container_width=True):
        bar = st.progress(0.0, text="Tarama hazırlanıyor...")
        with st.spinner("BIST hisseleri taranıyor..."):
            result = run_guclu_mum(symbols, lambda v, t: bar.progress(v, text=t), workers)
        bar.empty()
        st.session_state["guclu_mum_result"] = result

    result = st.session_state.get("guclu_mum_result")
    if result is None:
        return

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Günlük", len(result.gunluk))
    m2.metric("Haftalık", len(result.haftalik))
    m3.metric("Aylık", len(result.aylik))
    m4.metric("Takip", len(result.takip))

    tabs = st.tabs(["Günlük Güçlü Mum", "Haftalık Güçlü Mum", "Aylık Güçlü Mum", "Takip Listesi"])
    frames = [result.gunluk, result.haftalik, result.aylik, result.takip]
    for tab, frame in zip(tabs, frames):
        with tab:
            if frame.empty:
                st.warning("Bu bölümde sonuç yok.")
            else:
                st.dataframe(frame, use_container_width=True, hide_index=True, height=560)

    filename = f"bist_GUCLU_MUM_4_SAYFA_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    st.download_button(
        "4 Sayfalı Excel'i İndir",
        data=_excel_bytes(result),
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    if not result.failed.empty:
        with st.expander(f"Veri alınamayanlar ({len(result.failed)})"):
            st.dataframe(result.failed, use_container_width=True, hide_index=True)
